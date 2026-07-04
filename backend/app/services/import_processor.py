"""ImportProcessor — framework-independent Excel import execution.

Processes an uploaded Excel file in the background, creating Event entities
per validated row. Designed to run inside asyncio.create_task or a Celery worker
with zero code changes beyond the dispatch mechanism.
"""

import asyncio
from dataclasses import dataclass
from datetime import UTC, datetime
from logging import getLogger
from pathlib import Path
from uuid import UUID

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.config import settings
from app.models.enums import EventStatus, ImportRowError, ImportRowStatus, ImportStatus
from app.models.event import Event
from app.models.import_job import ImportJob
from app.models.import_row_result import ImportJobRowResult
from app.models.organizer import Organizer
from app.repositories.import_job import ImportJobRepository
from app.repositories.import_row_result import ImportJobRowResultRepository
from app.repositories.organizer import OrganizerRepository
from app.schemas.import_row import EventExcelRow

logger = getLogger(__name__)

_EXPECTED_HEADERS: set[str] = {
    "title",
    "description",
    "start_date",
    "location",
    "organizer_name",
}

_OPTIONAL_HEADERS: set[str] = {
    "short_description",
    "start_time",
    "end_time",
    "image_url",
    "registration_url",
    "category_id",
}


@dataclass(frozen=True, slots=True)
class ParsedRow:
    row_number: int
    title: str | None
    description: str | None
    start_date: str | None
    start_time: str | None
    end_time: str | None
    location: str | None
    organizer_name: str | None
    short_description: str | None
    image_url: str | None
    registration_url: str | None
    category_id: str | None


@dataclass(frozen=True, slots=True)
class RowResult:
    row_number: int
    status: ImportRowStatus
    created_entity_id: UUID | None = None
    error_code: ImportRowError | None = None
    error_message: str | None = None


class ImportProcessor:
    """Orchestrates Excel parsing, validation, batch event creation, and progress tracking.

    Framework-independent: receives primitives and a session factory.
    No FastAPI, no UploadFile, no Depends.
    """

    def __init__(
        self,
        session_factory: async_sessionmaker[AsyncSession],
    ) -> None:
        self._session_factory = session_factory

    async def process(self, job_id: UUID, file_path: Path, user_id: UUID) -> None:
        started = datetime.now(UTC)
        logger.info("Processing import job %s, file=%s", job_id, file_path)

        workbook = await _read_workbook(file_path)
        if workbook is None:
            await _fail_job(self._session_factory, job_id, "Corrupt or unreadable Excel file")
            return

        try:
            sheet = workbook.active
            headers = _extract_headers(sheet)
            missing = _EXPECTED_HEADERS - set(headers.keys())
            if missing:
                await _fail_job(
                    self._session_factory,
                    job_id,
                    f"Missing required columns: {', '.join(sorted(missing))}",
                )
                return

            total_rows = max(sheet.max_row - 1, 0)
            await _start_job(self._session_factory, job_id, total_rows)
            logger.info("Import job %s: total_rows=%d", job_id, total_rows)

            organizer_cache: dict[str, UUID | None] = {}
            results: list[RowResult] = []
            batch: list[tuple[Event, int, bool]] = []
            processed = 0

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                row_number = row_idx
                parsed = _parse_row(headers, row, row_number)
                if _is_empty_row(parsed):
                    continue

                outcome = await _validate_and_build_event(
                    self._session_factory,
                    parsed,
                    organizer_cache,
                    row_number,
                )

                if isinstance(outcome, RowResult):
                    results.append(outcome)
                    processed += 1
                elif isinstance(outcome, tuple):
                    event, used_default = outcome
                    batch.append((event, row_number, used_default))
                    if len(batch) >= settings.import_batch_size:
                        processed += await _flush_batch(
                            self._session_factory,
                            batch,
                            results,
                        )
                        batch.clear()
                        await _update_progress(self._session_factory, job_id, processed, results)

            if batch:
                processed += await _flush_batch(
                    self._session_factory,
                    batch,
                    results,
                )

            await _save_row_results(self._session_factory, job_id, results)
            await _complete_job(self._session_factory, job_id, results, started)

        finally:
            workbook.close()

        duration = int((datetime.now(UTC) - started).total_seconds())
        logger.info(
            "Import job %s completed in %ds",
            job_id,
            duration,
        )


async def _read_workbook(file_path: Path):
    try:
        return await asyncio.to_thread(openpyxl.load_workbook, str(file_path), read_only=True)
    except Exception:
        return None


def _extract_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(sheet[1], start=0):
        if cell.value is not None:
            name = str(cell.value).strip().lower()
            if name in _EXPECTED_HEADERS or name in _OPTIONAL_HEADERS:
                headers[name] = col_idx
    return headers


def _parse_row(headers: dict[str, int], row, row_number: int) -> ParsedRow:
    def _get(name: str) -> str | None:
        idx = headers.get(name)
        if idx is not None and idx < len(row):
            val = row[idx].value
            if val is None:
                return None
            stripped = str(val).strip()
            return stripped if stripped else None
        return None

    return ParsedRow(
        row_number=row_number,
        title=_get("title"),
        description=_get("description"),
        start_date=_get("start_date"),
        start_time=_get("start_time"),
        end_time=_get("end_time"),
        location=_get("location"),
        organizer_name=_get("organizer_name"),
        short_description=_get("short_description"),
        image_url=_get("image_url"),
        registration_url=_get("registration_url"),
        category_id=_get("category_id"),
    )


def _is_empty_row(parsed: ParsedRow) -> bool:
    return all(
        v is None
        for v in (
            parsed.title,
            parsed.description,
            parsed.start_date,
            parsed.location,
            parsed.organizer_name,
        )
    )


async def _find_organizer(
    session_factory: async_sessionmaker[AsyncSession],
    name: str,
    cache: dict[str, UUID | None],
) -> UUID | None:
    if name in cache:
        return cache[name]
    async with session_factory() as session:
        repo = OrganizerRepository(session)
        organizer: Organizer | None = await repo.get_by_name(name)
        result = organizer.id if organizer is not None else None
        cache[name] = result
        return result


def _resolve_category(row_category: UUID | None) -> tuple[UUID | None, bool]:
    if row_category is not None:
        return row_category, False
    if settings.default_import_category_id:
        try:
            return UUID(settings.default_import_category_id), True
        except ValueError:
            return None, False
    return None, False


async def _validate_and_build_event(
    session_factory: async_sessionmaker[AsyncSession],
    parsed: ParsedRow,
    organizer_cache: dict[str, UUID | None],
    row_number: int,
) -> "RowResult | tuple[Event, bool]":
    try:
        validated = EventExcelRow(
            title=parsed.title or "",
            description=parsed.description or "",
            start_date=parsed.start_date or "",
            start_time=parsed.start_time,
            end_time=parsed.end_time,
            location=parsed.location or "",
            organizer_name=parsed.organizer_name or "",
            short_description=parsed.short_description,
            image_url=parsed.image_url,
            registration_url=parsed.registration_url,
            category_id=UUID(parsed.category_id) if parsed.category_id else None,
        )
    except Exception as exc:
        return RowResult(
            row_number=row_number,
            status=ImportRowStatus.failed,
            error_code=ImportRowError.VALIDATION_ERROR,
            error_message=str(exc),
        )

    organizer_id = await _find_organizer(
        session_factory,
        validated.organizer_name,
        organizer_cache,
    )
    if organizer_id is None:
        return RowResult(
            row_number=row_number,
            status=ImportRowStatus.failed,
            error_code=ImportRowError.ORGANIZER_NOT_FOUND,
            error_message=f"Organizer '{validated.organizer_name}' not found",
        )

    category_id, used_default = _resolve_category(validated.category_id)
    if category_id is None:
        return RowResult(
            row_number=row_number,
            status=ImportRowStatus.failed,
            error_code=ImportRowError.CATEGORY_REQUIRED,
            error_message="No category_id in row and no default configured",
        )

    event = Event(
        title=validated.title,
        description=validated.description,
        short_description=validated.short_description,
        start_date=validated.start_date,
        start_time=validated.start_time,
        end_time=validated.end_time,
        location=validated.location,
        image_url=validated.image_url,
        registration_url=validated.registration_url,
        organizer_id=organizer_id,
        category_id=category_id,
        status=EventStatus.draft,
    )
    return event, used_default


async def _flush_batch(
    session_factory: async_sessionmaker[AsyncSession],
    batch: list[tuple[Event, int, bool]],
    results: list[RowResult],
) -> int:
    async with session_factory() as session:
        for evt, _, _ in batch:
            session.add(evt)

        try:
            await session.flush()
            await session.commit()
            for evt, row_number, used_default in batch:
                row_status = ImportRowStatus.warning if used_default else ImportRowStatus.imported
                results.append(
                    RowResult(
                        row_number=row_number,
                        status=row_status,
                        created_entity_id=evt.id,
                        error_code=ImportRowError.DEFAULT_CATEGORY_USED if used_default else None,
                        error_message="Default category_id used" if used_default else None,
                    )
                )
            return len(batch)
        except Exception:
            await session.rollback()
            return await _retry_batch_individually(
                session_factory,
                batch,
                results,
            )


async def _retry_batch_individually(
    session_factory: async_sessionmaker[AsyncSession],
    batch: list[tuple[Event, int, bool]],
    results: list[RowResult],
) -> int:
    count = 0
    for evt, row_number, used_default in batch:
        try:
            async with session_factory() as retry_session:
                retry_event = Event(
                    title=evt.title,
                    description=evt.description,
                    short_description=evt.short_description,
                    start_date=evt.start_date,
                    start_time=evt.start_time,
                    end_time=evt.end_time,
                    location=evt.location,
                    image_url=evt.image_url,
                    registration_url=evt.registration_url,
                    organizer_id=evt.organizer_id,
                    category_id=evt.category_id,
                    status=EventStatus.draft,
                )
                retry_session.add(retry_event)
                await retry_session.flush()
                await retry_session.commit()
                row_status = ImportRowStatus.warning if used_default else ImportRowStatus.imported
                results.append(
                    RowResult(
                        row_number=row_number,
                        status=row_status,
                        created_entity_id=retry_event.id,
                        error_code=ImportRowError.DEFAULT_CATEGORY_USED if used_default else None,
                        error_message="Default category_id used" if used_default else None,
                    )
                )
                count += 1
        except Exception as exc:
            results.append(
                RowResult(
                    row_number=row_number,
                    status=ImportRowStatus.failed,
                    error_code=ImportRowError.DATABASE_ERROR,
                    error_message=str(exc),
                )
            )
            count += 1
    return count


async def _fail_job(
    session_factory: async_sessionmaker[AsyncSession],
    job_id: UUID,
    error_message: str,
) -> None:
    logger.error("Import job %s failed: %s", job_id, error_message)
    async with session_factory() as session:
        repo = ImportJobRepository(session)
        job: ImportJob | None = await repo.get_by_id(job_id)
        if job is not None:
            job.status = ImportStatus.failed
            job.error_message = error_message
            job.finished_at = datetime.now(UTC)
            await session.commit()


async def _start_job(
    session_factory: async_sessionmaker[AsyncSession],
    job_id: UUID,
    total_rows: int,
) -> None:
    async with session_factory() as session:
        repo = ImportJobRepository(session)
        job: ImportJob | None = await repo.get_by_id(job_id)
        if job is not None:
            job.total_rows = total_rows
            job.started_at = datetime.now(UTC)
            await session.commit()


async def _update_progress(
    session_factory: async_sessionmaker[AsyncSession],
    job_id: UUID,
    processed: int,
    results: list[RowResult],
) -> None:
    imported = sum(1 for r in results if r.status == ImportRowStatus.imported)
    warning = sum(1 for r in results if r.status == ImportRowStatus.warning)
    failed = sum(1 for r in results if r.status == ImportRowStatus.failed)
    async with session_factory() as session:
        repo = ImportJobRepository(session)
        job: ImportJob | None = await repo.get_by_id(job_id)
        if job is not None:
            job.processed_rows = processed
            job.imported_rows = imported
            job.warning_rows = warning
            job.failed_rows = failed
            await session.commit()


async def _complete_job(
    session_factory: async_sessionmaker[AsyncSession],
    job_id: UUID,
    results: list[RowResult],
    started: datetime,
) -> None:
    imported = sum(1 for r in results if r.status == ImportRowStatus.imported)
    warning = sum(1 for r in results if r.status == ImportRowStatus.warning)
    failed = sum(1 for r in results if r.status == ImportRowStatus.failed)
    duration = int((datetime.now(UTC) - started).total_seconds())
    async with session_factory() as session:
        repo = ImportJobRepository(session)
        job: ImportJob | None = await repo.get_by_id(job_id)
        if job is not None:
            job.status = ImportStatus.completed
            job.imported_rows = imported
            job.warning_rows = warning
            job.failed_rows = failed
            job.processed_rows = imported + warning + failed
            job.duration = duration
            job.finished_at = datetime.now(UTC)
            await session.commit()


async def _save_row_results(
    session_factory: async_sessionmaker[AsyncSession],
    job_id: UUID,
    results: list[RowResult],
) -> None:
    if not results:
        return
    async with session_factory() as session:
        repo = ImportJobRowResultRepository(session)
        entities: list[ImportJobRowResult] = []
        for r in results:
            entity = ImportJobRowResult(
                import_job_id=job_id,
                row_number=r.row_number,
                status=r.status,
                created_entity_id=r.created_entity_id,
                error_code=r.error_code.value if r.error_code is not None else None,
                error_message=r.error_message,
            )
            entities.append(entity)
        await repo.create_many(entities)
        await session.commit()
